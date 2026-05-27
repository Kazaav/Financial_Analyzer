"""Tests for export module."""
from __future__ import annotations

import io
import json

from openpyxl import load_workbook

from app.export import build_filename, to_csv, to_json, to_xlsx
from app.models import AnalysisRecord


def _make_payload():
    """Build a minimal payload structure matching build_export_payload output."""
    return {
        "analysis_id": "test-1",
        "created_at": "2026-05-27 10:00",
        "mode": "same_year",
        "mode_label": "多社同年度比較",
        "selected_year": 2024,
        "selected_company": None,
        "kpis": {
            "document_count": 2,
            "company_count": 2,
            "year_range": "2024",
            "avg_confidence": 1.0,
            "total_revenue": 300_000,
            "total_net_income": 30_000,
        },
        "rows": [
            {
                "doc_id": "d1", "filename": "a.pdf",
                "company_name": "Test A", "security_code": "0001",
                "edinet_code": "E001", "fiscal_year": 2024,
                "fiscal_period": "", "unit": "百万円", "confidence": 1.0,
                "raw__revenue": 100_000, "raw__net_income": 10_000,
                "derived__roa": 0.05, "growth__revenue_growth": 0.1,
            },
            {
                "doc_id": "d2", "filename": "b.pdf",
                "company_name": "Test B", "security_code": "0002",
                "edinet_code": "E002", "fiscal_year": 2024,
                "fiscal_period": "", "unit": "百万円", "confidence": 1.0,
                "raw__revenue": 200_000, "raw__net_income": 20_000,
                "derived__roa": 0.04, "growth__revenue_growth": 0.05,
            },
        ],
        "metric_sources": {
            "d1": {"revenue": {"label": "売上高", "section": "主要な経営指標等の推移", "page": 2, "raw_value": 100_000}},
        },
        "formulas": {"roa": "当期純利益 ÷ 総資産"},
        "raw_labels": {"revenue": "売上高", "net_income": "当期純利益"},
        "derived_labels": {"roa": "ROA"},
        "growth_labels": {"revenue_growth": "売上成長率"},
    }


class TestCsv:
    def test_csv_has_bom(self):
        result = to_csv(_make_payload())
        # UTF-8 BOM ensures Excel opens it correctly with Japanese text
        assert result.startswith("﻿")

    def test_csv_has_header_and_rows(self):
        result = to_csv(_make_payload())
        lines = result.lstrip("﻿").strip().split("\n")
        assert len(lines) == 3  # 1 header + 2 rows
        header = lines[0]
        assert "doc_id" in header
        assert "raw__revenue" in header

    def test_csv_handles_none(self):
        payload = _make_payload()
        payload["rows"][0]["raw__net_income"] = None
        result = to_csv(payload)
        # None should become empty string, not literal 'None'
        assert ",None," not in result


class TestJson:
    def test_json_roundtrip(self):
        result = to_json(_make_payload())
        decoded = json.loads(result)
        assert decoded["analysis_id"] == "test-1"
        assert len(decoded["rows"]) == 2
        assert decoded["rows"][0]["raw__revenue"] == 100_000

    def test_json_japanese_not_escaped(self):
        result = to_json(_make_payload())
        # ensure_ascii=False keeps 売上高 readable
        assert "売上高" in result


class TestXlsx:
    def test_xlsx_returns_bytes(self):
        body = to_xlsx(_make_payload())
        # XLSX is a ZIP container — should start with PK\x03\x04
        assert body[:2] == b"PK"

    def test_xlsx_has_three_sheets(self):
        body = to_xlsx(_make_payload())
        wb = load_workbook(io.BytesIO(body))
        assert wb.sheetnames == ["主要指標", "抽出ソース", "派生指標 計算式"]

    def test_xlsx_main_sheet_data(self):
        body = to_xlsx(_make_payload())
        wb = load_workbook(io.BytesIO(body))
        ws = wb["主要指標"]
        # Header row exists, then 2 data rows
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3
        # Verify header includes the Japanese label
        assert "売上高" in rows[0]


class TestFilename:
    def test_filename_sanitised(self):
        record = AnalysisRecord(id="demo-it-services", created_at="2026", documents=[])
        assert build_filename(record, "same_year", "csv") == "financial-analysis_demo-it-services_same_year.csv"
