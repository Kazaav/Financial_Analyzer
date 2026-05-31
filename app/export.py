"""Export an analysis as CSV / Excel / JSON.

Each export contains the same logical content:
  - Document-level rows: company / year / unit / raw metrics / derived metrics / growth
  - Each metric column carries its source (label, section, page) for traceability
"""
from __future__ import annotations

import csv
import io
import json
from typing import Any

from .analysis import DERIVED_LABELS, FORMULAS, METRIC_ORDER
from .models import AnalysisRecord
from .pdf_parser import DISPLAY_NAMES

GROWTH_LABELS = {
    "revenue_growth": "売上成長率",
    "net_income_growth": "純利益成長率",
}


def _build_rows(record: AnalysisRecord, analysis: dict[str, Any]) -> list[dict[str, Any]]:
    """Reshape analysis rows into a flat export structure."""
    out: list[dict[str, Any]] = []
    for row in analysis["rows"]:
        doc = row["doc"]
        entry: dict[str, Any] = {
            "doc_id": doc.id,
            "filename": doc.filename,
            "company_name": doc.company_name,
            "security_code": doc.security_code,
            "edinet_code": doc.edinet_code,
            "fiscal_year": doc.fiscal_year,
            "fiscal_period": doc.fiscal_period,
            "unit": doc.unit,
            "accounting_standard": getattr(doc, "accounting_standard", "JGAAP"),
            "confidence": doc.confidence,
        }
        # Raw metrics
        for key in METRIC_ORDER:
            entry[f"raw__{key}"] = row["metrics"].get(key)
        # Derived metrics
        for key in DERIVED_LABELS:
            entry[f"derived__{key}"] = row.get("derived", {}).get(key)
        # Growth metrics
        for key in GROWTH_LABELS:
            entry[f"growth__{key}"] = row.get("growth", {}).get(key)
        out.append(entry)
    return out


def build_export_payload(record: AnalysisRecord, analysis: dict[str, Any]) -> dict[str, Any]:
    rows = _build_rows(record, analysis)
    # Build a metric-source map: doc_id -> { metric_key: {label, section, page} }
    sources: dict[str, dict[str, Any]] = {}
    for doc in record.documents:
        if doc.metric_sources:
            sources[doc.id] = doc.metric_sources

    return {
        "analysis_id": record.id,
        "created_at": record.created_at,
        "mode": analysis.get("mode"),
        "mode_label": analysis.get("mode_label"),
        "selected_year": analysis.get("selected_year"),
        "selected_company": analysis.get("selected_company"),
        "kpis": {
            "document_count": analysis["kpis"]["document_count"],
            "company_count": analysis["kpis"]["company_count"],
            "year_range": analysis["kpis"]["year_range"],
            "avg_confidence": analysis["kpis"]["avg_confidence"],
            "total_revenue": analysis["kpis"]["total_revenue"],
            "total_net_income": analysis["kpis"]["total_net_income"],
        },
        "rows": rows,
        "metric_sources": sources,
        "formulas": {k: FORMULAS[k] for k in FORMULAS if k in FORMULAS},
        "raw_labels": DISPLAY_NAMES,
        "derived_labels": DERIVED_LABELS,
        "growth_labels": GROWTH_LABELS,
    }


# ─── CSV ───────────────────────────────────────────────────────────
def to_csv(payload: dict[str, Any]) -> str:
    rows = payload["rows"]
    if not rows:
        return ""
    buf = io.StringIO()
    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    for r in rows:
        writer.writerow({k: ("" if v is None else v) for k, v in r.items()})
    # Prepend UTF-8 BOM so Excel opens it correctly with Japanese text
    return "﻿" + buf.getvalue()


# ─── JSON ──────────────────────────────────────────────────────────
def to_json(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2, default=str)


# ─── XLSX ──────────────────────────────────────────────────────────
def to_xlsx(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Default sheet renamed
    main_sheet = wb.active
    assert main_sheet is not None
    main_sheet.title = "主要指標"

    # Build pretty header labels combining the prefix and Japanese label
    raw_labels = payload["raw_labels"]
    derived_labels = payload["derived_labels"]
    growth_labels = payload["growth_labels"]

    META_COLUMNS = [
        ("doc_id", "Doc ID"),
        ("filename", "ファイル名"),
        ("security_code", "証券コード"),
        ("edinet_code", "EDINETコード"),
        ("company_name", "企業名"),
        ("fiscal_year", "事業年度"),
        ("fiscal_period", "事業期間"),
        ("unit", "単位"),
        ("accounting_standard", "会計基準"),
        ("confidence", "抽出信頼度"),
    ]
    raw_cols = [(f"raw__{k}", raw_labels.get(k, k)) for k in raw_labels]
    derived_cols = [(f"derived__{k}", derived_labels[k]) for k in derived_labels]
    growth_cols = [(f"growth__{k}", growth_labels[k]) for k in growth_labels]
    all_cols = META_COLUMNS + raw_cols + derived_cols + growth_cols

    header_fill = PatternFill("solid", fgColor="11151E")
    header_font = Font(color="E6E8EE", bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center")

    # Header
    for col_idx, (_, label) in enumerate(all_cols, start=1):
        c = main_sheet.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    # Rows
    for row_idx, row in enumerate(payload["rows"], start=2):
        for col_idx, (key, _) in enumerate(all_cols, start=1):
            value = row.get(key)
            main_sheet.cell(row=row_idx, column=col_idx, value=value)

    # Freeze first column + header
    main_sheet.freeze_panes = "C2"
    # Column widths (rough auto)
    for col_idx, (_key, label) in enumerate(all_cols, start=1):
        width = max(12, min(28, len(label) * 2.2 + 4))
        main_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # ─── Sheet 2: 抽出ソース ─────────────────────────────────────────
    src_sheet = wb.create_sheet("抽出ソース")
    src_header = ["Doc ID", "ファイル名", "指標", "抽出ラベル", "セクション", "ページ", "値"]
    for col_idx, label in enumerate(src_header, start=1):
        c = src_sheet.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    row_idx = 2
    doc_map = {r["doc_id"]: r for r in payload["rows"]}
    for doc_id, metric_map in payload["metric_sources"].items():
        doc_row = doc_map.get(doc_id, {})
        for metric_key, info in metric_map.items():
            src_sheet.cell(row=row_idx, column=1, value=doc_id)
            src_sheet.cell(row=row_idx, column=2, value=doc_row.get("filename"))
            src_sheet.cell(row=row_idx, column=3, value=raw_labels.get(metric_key, metric_key))
            src_sheet.cell(row=row_idx, column=4, value=info.get("label"))
            src_sheet.cell(row=row_idx, column=5, value=info.get("section"))
            src_sheet.cell(row=row_idx, column=6, value=info.get("page"))
            src_sheet.cell(row=row_idx, column=7, value=info.get("raw_value"))
            row_idx += 1
    for col_idx in range(1, len(src_header) + 1):
        src_sheet.column_dimensions[get_column_letter(col_idx)].width = 18
    src_sheet.freeze_panes = "A2"

    # ─── Sheet 3: 計算式 ────────────────────────────────────────────
    formula_sheet = wb.create_sheet("派生指標 計算式")
    for col_idx, label in enumerate(["指標キー", "ラベル", "計算式"], start=1):
        c = formula_sheet.cell(row=1, column=col_idx, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
    row_idx = 2
    for key, formula in payload["formulas"].items():
        formula_sheet.cell(row=row_idx, column=1, value=key)
        formula_sheet.cell(row=row_idx, column=2, value=derived_labels.get(key) or growth_labels.get(key) or key)
        formula_sheet.cell(row=row_idx, column=3, value=formula)
        row_idx += 1
    formula_sheet.column_dimensions["A"].width = 28
    formula_sheet.column_dimensions["B"].width = 26
    formula_sheet.column_dimensions["C"].width = 60
    formula_sheet.freeze_panes = "A2"

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_filename(record: AnalysisRecord, mode: str, ext: str) -> str:
    """Sanitised filename for download Content-Disposition."""
    safe_id = record.id.replace("/", "_")
    safe_mode = mode.replace("/", "_")
    return f"financial-analysis_{safe_id}_{safe_mode}.{ext}"
